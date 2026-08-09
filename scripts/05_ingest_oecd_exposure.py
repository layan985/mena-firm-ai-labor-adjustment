from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data/raw/oecd_ai_exposure/oecd_ai_exposure_2026.xlsx"
CSV_OUT = ROOT / "data/processed/oecd_occupation_ai_exposure.csv"
PARQUET_OUT = ROOT / "data/processed/oecd_occupation_ai_exposure.parquet"
AUDIT = ROOT / "outputs/audits/oecd_exposure_summary.json"

COLS = {
    "OCC_Code": "occupation_code",
    "OCC_Title": "occupation_title",
    "AI Capability Gap Index_Rev. norm.": "ai_exposure",
    "AI Capability Gap Index_Total": "ai_capability_gap_total",
}


def main() -> None:
    if not RAW.exists():
        raise SystemExit("Missing OECD workbook. Run: python scripts/04_fetch_oecd_exposure.py")
    wb = load_workbook(RAW, data_only=True, read_only=True)
    ws = wb["Data"]
    headers = [c.value for c in next(ws.iter_rows(min_row=3, max_row=3))]
    idx = {name: headers.index(name) for name in COLS}
    records = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[idx["OCC_Code"]] is None:
            continue
        records.append({new: row[idx[old]] for old, new in COLS.items()})

    CSV_OUT.parent.mkdir(parents=True, exist_ok=True)
    with CSV_OUT.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(COLS.values()))
        writer.writeheader()
        writer.writerows(records)

    parquet_written = False
    try:
        import duckdb
        con = duckdb.connect()
        con.execute(
            "COPY (SELECT * FROM read_csv_auto(?)) TO ? (FORMAT PARQUET, COMPRESSION ZSTD)",
            [str(CSV_OUT), str(PARQUET_OUT)],
        )
        con.close()
        parquet_written = True
    except ImportError:
        pass

    exposures = [float(r["ai_exposure"]) for r in records]
    summary = {
        "rows": len(records),
        "unique_occupation_codes": len({r["occupation_code"] for r in records}),
        "ai_exposure_min": min(exposures),
        "ai_exposure_mean": sum(exposures) / len(exposures),
        "ai_exposure_max": max(exposures),
        "csv_written": str(CSV_OUT.relative_to(ROOT)),
        "parquet_written": parquet_written,
    }
    AUDIT.parent.mkdir(parents=True, exist_ok=True)
    AUDIT.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
